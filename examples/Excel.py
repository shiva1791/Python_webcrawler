__author__ = 'Shiva'
import openpyxl
import copy


class excel:
    wb = openpyxl.load_workbook('Defect.xlsx')
    sheetnames = wb.get_sheet_names()  # to get sheet names in the excel

    def getSprintDetails(self):

        sheet = excel.wb.get_sheet_by_name(excel.sheetnames[1])
        last_row = sheet.max_row  # last row in the sheet
        sprintname = []
        sprintstartdate = []
        sprintenddate = []
        sprintdetail = []
        for i in range(2, last_row + 1, 1):
            sprintname.append(sheet.cell(row=i, column=1).value)
            sprintstartdate.append(sheet.cell(row=i, column=2).value)
            sprintenddate.append(sheet.cell(row=i, column=3).value)

        for i in range(len(sprintname)):
            sprintdetail.append([sprintname[i], sprintstartdate[i], sprintenddate[i]])
        return (sprintdetail)

    def getDefectDetails(self):

        sheet = excel.wb.get_sheet_by_name(excel.sheetnames[0])
        last_row = sheet.max_row
        #print(last_row)
        defect_date = []
        defect_priority = []
        defect_deatil = []
        for i in range(2, last_row + 1):
            defect_date.append(sheet.cell(row=i, column=2).value)
            defect_priority.append(sheet.cell(row=i, column=3).value)
        for i in range(len(defect_date)):
            defect_deatil.append([defect_date[i], defect_priority[i]])
            pass
        return (defect_deatil)

    def getDefectInSprint(self):
        sprintdetails = excel.getSprintDetails(self)
        defectdetails = excel.getDefectDetails(self)

        sheet = excel.wb.get_sheet_by_name(excel.sheetnames[0])
        last_row = sheet.max_row
        #print(last_row)

        # add month and year column in defect list
        #print(len(defectdetails))
        for i in range(len(defectdetails)):
            for j in range(len(sprintdetails)):
                if (defectdetails[i][0] >= sprintdetails[j][1] and defectdetails[i][0] <= sprintdetails[j][2]):
                    # print(sprintdetails[j][0])
                    value = [sprintdetails[j][0]]
                    # print(value)
                    defectdetails[i].extend(value)
        months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October',
                  'November', 'December']
        monthly_defect = []
        month_year_name = []
        defect_count_month = []
        priority_1 = []
        priority_2 = []
        priority_3 = []
        for i in range(len(defectdetails)):
            for year in range(2015, 2017, 1):
                for month_val in range(1, 13, 1):
                    if defectdetails[i][0].month == month_val and defectdetails[i][0].year == year:
                        month_year_val = [str((str(months[month_val - 1]) + ' ' + str(year)))]
                        defectdetails[i].extend(month_year_val)
                        pass
        #excel.writeinsheet(self, defectdetails)

        for i in range(len(defectdetails)):
            # print(defectdetails[i])
            pass
        #print('----------------------------------------------------------------------------------------------')

        # For counting defects monthly based on priority
        for year in range(2015, 2017, 1):
            for month_val in range(1, 13, 1):
                value_to_search = str((str(months[month_val - 1]) + ' ' + str(year)))
                count = 0
                p1, p2, p3 = 0, 0, 0
                #print('-------')
                for i in range(len(defectdetails)):
                    if value_to_search in defectdetails[i]:
                        count += 1
                        #print(count)
                        if defectdetails[i][1] == 1:
                            p1 += 1
                        elif defectdetails[i][1] == 2:
                            p2 += 1
                        elif defectdetails[i][1] == 3:
                            p3 += 1
                        else:
                            pass
                if count > 0:
                    month_year_name.append(value_to_search)
                    defect_count_month.append(count)
                if p1 >= 0 or p2 >= 0 or p3 >= 0:
                    priority_1.append(p1)
                    priority_2.append(p2)
                    priority_3.append(p3)
            pass

        for i in range(len(month_year_name)):
            monthly_defect.append(
                [month_year_name[i], defect_count_month[i], priority_1[i], priority_2[i], priority_3[i]])

        for i in range(len(monthly_defect)):
            pass
            # print(monthly_defect[i])
        #excel.writeinsheet(self, monthly_defect)

        # For counting defects based on sprint
        # this will determine the number of defects each sprint for total defect and by priority

        for i in range(len(defectdetails)):
            for j in range(len(sprintdetails)):
                if (sprintdetails[j][0] == defectdetails[i][2]):
                    if (len(sprintdetails[j]) == 3):
                        sprintdetails[j].append(0)  # for total defect
                        sprintdetails[j].append(0)  # for p1 defect
                        sprintdetails[j].append(0)  # for p2 defect
                        sprintdetails[j].append(0)  # for p3 defect
                    sprintdetails[j][3] += 1

                    if defectdetails[i][1] == 1:
                        sprintdetails[j][4] += 1

                    elif defectdetails[i][1] == 2:
                        sprintdetails[j][5] += 1

                    elif defectdetails[i][1] == 3:
                        sprintdetails[j][6] += 1
        #excel.writeinsheet(self, sprintdetails)
        return(defectdetails)

    def getDefectsPerDate(self):
        defectdetails = excel.getDefectInSprint(self)
        defectdetailscopy = list(defectdetails)

        print(len(defectdetailscopy))
        print('----------------')
        print(defectdetails)
        Date_sort= []
        for i in range(len(defectdetails)):
            value = len(defectdetailscopy)
            for j in range(0,value,1):
                print(j)
                if (defectdetails[i][0] == defectdetailscopy[j][0]):
                    Date_sort.append(str(defectdetails[i][1]))
                    #defectdetailscopy.pop(j)
                    value = len(defectdetailscopy)
                    #value -=1
                    print(str(value)+'is the new length of the copy list')

        print(Date_sort)


        pass

    def getDefectDetail_Dictionary(self):
        sheet = excel.wb.get_sheet_by_name(excel.sheetnames[0])
        _dict_defect_data = {}
        last_row = sheet.max_row
        print(last_row)

        for rows in range(1,last_row+1,1):
            defectdate = sheet.cell(row = rows,column=2).value

        pass

    def writeinsheet(self, anylist):
        anylist = anylist
        wb = openpyxl.load_workbook('Defect.xlsx')
        sheetnames = wb.get_sheet_names()  # to get sheet names in the excel
        sheet = wb.get_sheet_by_name(sheetnames[2])
        #wb.remove_sheet(sheet)

        new_sheet = wb.create_sheet()
        wb.save('Defect.xlsx')
        #sheet = wb.get_sheet_by_name('Temp')

        for rows in range(len(anylist)):
            for columns in range(len(anylist[rows])):
                new_sheet.cell(row=rows + 1, column=columns + 1, value=str(anylist[rows][columns]))
                # print(anylist[rows][columns])
        wb.save('Defect.xlsx')


if __name__ == '__main__':
    e = excel()
    # e.getSprintDetails()
    # e.getDefectDetails()
    #e.getDefectInSprint()
   #e.getDefectsPerDate()
    e.getDefectDetail_Dictionary()
